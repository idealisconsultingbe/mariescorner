# -*- coding: utf-8 -*-
"""
Archive unused attribute variants.
"""

import argparse
import logging
from openpyxl import load_workbook


def read_csv(attributes_xlsx):
    attribute_combinaison = {}
    wb = load_workbook(attributes_xlsx)
    sheet = wb.active
    for row in range(2, end_row + 1):
        cell_main_attribute = sheet.cell(row=row, column=1)
        main_attribute_value = cell_main_attribute.value
        attribute_combinaison[main_attribute_value] = []
        column = 2
        cell_secondary_attribute = sheet.cell(row=row, column=column)
        secondary_attribute_value = cell_secondary_attribute.value
        while secondary_attribute_value:
            attribute_combinaison[main_attribute_value].append(secondary_attribute_value)
            column += 1
            cell_secondary_attribute = sheet.cell(row=row, column=column)
            secondary_attribute_value = cell_secondary_attribute.value
    return attribute_combinaison


parser = argparse.ArgumentParser(description="")
parser.add_argument('database')
parser.add_argument('product_template_id')
parser.add_argument('main_attribute_id')
parser.add_argument('secondary_attribute_id')
parser.add_argument('xlsx_file')
args = parser.parse_args()

end_row = int(input("What is the last line of your xlsx file? "))
session.open(db=args.database)
attribute_combinaison = read_csv(args.xlsx_file)
product_tmpl = session.env['product.template'].browse(int(args.product_template_id))
main_attribute_id = int(args.main_attribute_id)
secondary_attribute_id = int(args.secondary_attribute_id)

logging.info('Start deleting unnecessary variants.')
created_variants = 0
for main_attribute_value in attribute_combinaison:
    product_template_main_attribute = session.env['product.template.attribute.value'].search([('product_attribute_value_id.name', '=', main_attribute_value), ('product_tmpl_id', '=', product_tmpl.id), ('attribute_id', '=', main_attribute_id)])
    product_template_secondary_attribute = session.env['product.template.attribute.value'].search([('product_attribute_value_id.name', 'in', attribute_combinaison[main_attribute_value]), ('product_tmpl_id', '=', product_tmpl.id), ('attribute_id', '=', secondary_attribute_id)])
    for ptsa in product_template_secondary_attribute:
        product_tmpl._create_product_variant((product_template_main_attribute + ptsa))
        created_variants += 1
    session.cr.commit()
    logging.info('#%s variants created' % created_variants)
logging.info('Nnecessary product variants created')

session.cr.commit()
session.close()
logging.info('Done')

