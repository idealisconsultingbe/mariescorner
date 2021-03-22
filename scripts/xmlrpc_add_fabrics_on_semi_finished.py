import xmlrpc.client
import argparse
from openpyxl import load_workbook


def read_csv(attributes_xlsx):
    """
    Read an xlsx file, should contain the following column
    Row 1 = Header
    * Fabrics        * Colors   *           *           * ...
    ******************************************************************************
    * fabric_name1   * Color1   * Color2    * Color 3   * ...
    * fabric_name2   * Color1   * Color2    *           * ...
    return: dict{fabric_name: [color1, color2, ...]}
    """
    attribute_combinaison = {}
    wb = load_workbook(attributes_xlsx)
    sheet = wb.active
    row = 2
    while sheet.cell(row=row, column=1).value:
        cell_fabric_attribute = sheet.cell(row=row, column=1)
        fabric_attribute_value = cell_fabric_attribute.value
        attribute_combinaison[fabric_attribute_value] = []
        column = 2
        cell_color_attribute = sheet.cell(row=row, column=column)
        color_attribute_value = cell_color_attribute.value
        while color_attribute_value:
            attribute_combinaison[fabric_attribute_value].append(str(color_attribute_value))
            column += 1
            cell_color_attribute = sheet.cell(row=row, column=column)
            color_attribute_value = cell_color_attribute.value
        row += 1
    return attribute_combinaison


def inverse_attribute_combinaison(attribute_combinaison):
    """
    Inverse the key, values of a dictionary
    :param attribute_combinaison: dict{fabric_name: [color1, color2, ...]}
    :return: dict{color_name: [fabric1, fabric2, ...]}
    """
    attribute_combinaison_inverse = {}
    for key, values in attribute_combinaison.items():
        for value in values:
            if value not in attribute_combinaison_inverse:
                attribute_combinaison_inverse[value] = [key]
            else:
                attribute_combinaison_inverse[value].append(key)
    return attribute_combinaison_inverse


def add_pav_on_products(models, product_ids, attribute_id, attribute_value_ids):
    """
    Write on product.template.attribute.line
    """
    ptal_ids = models.execute_kw(db, uid, password,
                                 'product.template.attribute.line',
                                 'search',
                                 [[
                                     ('product_tmpl_id', 'in', product_ids),
                                     ('attribute_id', '=', attribute_id),
                                 ]])
    models.execute_kw(db, uid, password, 'product.template.attribute.line', 'write', [ptal_ids, {
        'value_ids': [(6, 0, attribute_value_ids)],
    }])


def sort_product_ids_by_tcl(product_ids):
    """
    :param product_ids: list of dictionary of product.template
    :return: product_ids split by 'tcl' or 'no tcl'
    """
    no_tcl_ids = []
    tcl_ids = []
    for product in product_ids:
        if 'c.o.m.' in product["name"].lower() or 'com' in product["name"].lower() or (product["default_code"] and 'tcl' in product["default_code"].lower()):
            tcl_ids.append(product["id"])
        else:
            no_tcl_ids.append(product["id"])
    return no_tcl_ids, tcl_ids


def get_attribute_values(models, attribute_id, values_name, tcl=False):
    domain = [
        ('name', 'in', values_name),
        ('attribute_id', '=', attribute_id),
    ]
    if not tcl:
        domain.append(('name', '!=', 'TCL'),)
    pav_ids = models.execute_kw(db, uid, password,
                                'product.attribute.value',
                                'search',
                                [domain])
    return pav_ids


def get_tcl_attribute_values(models, attribute_id):
    domain = [
        ('attribute_id', '=', attribute_id),
        '|',
        ('name', 'like', '%DEFINIR GRADE%'),
        ('name', 'in', ['None', 'TCL']),
    ]
    pav_ids = models.execute_kw(db, uid, password,
                                'product.attribute.value',
                                'search',
                                [domain])
    return pav_ids


def get_dict_records(records_list):
    records = {}
    for record_list in records_list:
        records[record_list['id']] = record_list
    return records


def get_unallowed_ptav_fabric_by_product(ptav_list):
    unallowed_ptav_fabric_by_product = {}
    for ptav in ptav_list:
        product_tmpl_id = ptav['product_tmpl_id'][0]
        if product_tmpl_id not in unallowed_ptav_fabric_by_product:
            unallowed_ptav_fabric_by_product[product_tmpl_id] = [ptav['id']]
        else:
            unallowed_ptav_fabric_by_product[product_tmpl_id].append(ptav['id'])
    return unallowed_ptav_fabric_by_product


def configure_ptav_exclusion(color_name, color_attribute_ids, fabric_attribute_ids, product_ids):
    exclusion_to_create = []
    for attribute_color_id, attribute_fabric_id in zip(color_attribute_ids, fabric_attribute_ids):
        ptav_color_infos = models.execute_kw(db, uid, password,
                                             'product.template.attribute.value',
                                             'search_read',
                                             [[
                                                 ('name', '=', color_name),
                                                 ('attribute_id', '=', attribute_color_id),
                                                 ('product_tmpl_id', 'in', product_ids)
                                             ]],
                                             {'fields': ['product_tmpl_id']}
                                             )
        ptav_color_ids = [ptav["id"] for ptav in ptav_color_infos]
        ptav_color_records = get_dict_records(ptav_color_infos)
        unallowed_ptav_fabric_infos = models.execute_kw(db, uid, password,
                                                        'product.template.attribute.value',
                                                        'search_read',
                                                        [[
                                                            ('name', 'not in', attribute_combinaison_inverse[color_attribute_value]),
                                                            ('product_tmpl_id', 'in', product_ids),
                                                            ('attribute_id', '=', attribute_fabric_id)
                                                        ]],
                                                        {'fields': ['product_tmpl_id']}
                                                        )
        unallowed_ptav_fabric_ids = [ptav["id"] for ptav in unallowed_ptav_fabric_infos]
        unallowed_ptav_fabric_by_product = get_unallowed_ptav_fabric_by_product(unallowed_ptav_fabric_infos)
        if ptav_color_ids and unallowed_ptav_fabric_ids:
            ptae_infos = models.execute_kw(db, uid, password,
                                           'product.template.attribute.exclusion',
                                           'search_read',
                                           [[
                                               ('product_template_attribute_value_id', 'in', ptav_color_ids),
                                               ('product_tmpl_id', '=', product_ids),
                                           ]],
                                           {'fields': ['product_template_attribute_value_id', 'product_tmpl_id', 'value_ids']}
                                           )
            ptae_ids = [ptae["id"] for ptae in ptae_infos]
            ptae_records = get_dict_records(ptae_infos)
            ptav_color_with_exclusion_ids = [ptae["product_template_attribute_value_id"][0] for ptae in ptae_infos]
            ptav_color_without_exclusion_ids = [ptav for ptav in ptav_color_ids if ptav not in ptav_color_with_exclusion_ids]
            if ptae_ids:
                for ptae_id in ptae_records:
                    product_tmpl_id = ptae_records[ptae_id]['product_tmpl_id'][0]
                    excluded_values = ptae_records[ptae_id]['value_ids']
                    values_to_exclude = [(4, value_id) for value_id in unallowed_ptav_fabric_by_product[product_tmpl_id] if value_id not in excluded_values]
                    values_to_unexclude = [(3, value_id) for value_id in excluded_values if value_id not in unallowed_ptav_fabric_by_product[product_tmpl_id]]
                    if values_to_exclude or values_to_unexclude:
                        models.execute_kw(db, uid, password, 'product.template.attribute.exclusion', 'write', [ptae_id, {
                            'value_ids': values_to_exclude + values_to_unexclude,
                        }])
            for ptav_color_id in ptav_color_without_exclusion_ids:
                product_tmpl_id = ptav_color_records[ptav_color_id]['product_tmpl_id'][0]
                exclusion_to_create.append({'product_template_attribute_value_id': ptav_color_id,
                                            'product_tmpl_id': product_tmpl_id,
                                            'value_ids': [(6, 0, unallowed_ptav_fabric_by_product[product_tmpl_id])]})
    return exclusion_to_create


parser = argparse.ArgumentParser(description="Add Fabrics on semi finished")
parser.add_argument('url')
parser.add_argument('database')
parser.add_argument('user')
parser.add_argument('password')
parser.add_argument('product_category_id')
parser.add_argument('main_fabric_attribute_id')
parser.add_argument('main_color_attribute_id')
parser.add_argument('secondary_fabric_attribute_id')
parser.add_argument('secondary_color_attribute_id')
parser.add_argument('product_fabric_id')
parser.add_argument('fabric_fabric_attribute_id')
parser.add_argument('fabric_color_attribute_id')
parser.add_argument('xlsx_file')
args = parser.parse_args()

db =args.database
password = args.password
seat_main_fabric_id = int(args.main_fabric_attribute_id)
seat_main_color_id = int(args.main_color_attribute_id)
seat_secondary_fabric_id = int(args.secondary_fabric_attribute_id)
seat_secondary_color_id = int(args.secondary_color_attribute_id)
fabric_product_id = int(args.product_fabric_id)
fabric_main_fabric_id = int(args.fabric_fabric_attribute_id)
fabric_main_color_id = int(args.fabric_color_attribute_id)

common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(args.url))
uid = common.authenticate(db, args.user, password, {})
models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(args.url))

print('Load product category.')
product_category_ids = models.execute_kw(db, uid, password, 'product.category', 'search', [[('id', 'child_of', int(args.product_category_id))]])
print('%s Product category loaded.' % len(product_category_ids))

print('Load product templates.')
products_info = models.execute_kw(db, uid, password, 'product.template', 'search_read', [[('categ_id', 'in', product_category_ids)]], {'fields': ['name', 'default_code']})
print('%s Product templates loaded.' % len(products_info))

print('Read xlsx.')
attribute_combinaison = read_csv(args.xlsx_file)
attribute_combinaison_inverse = inverse_attribute_combinaison(attribute_combinaison)

print('Set Data1.')
fabrics_name = list(attribute_combinaison.keys())
color_name = list(attribute_combinaison_inverse.keys())
product_no_tcl_ids, tcl_ids = sort_product_ids_by_tcl(products_info)
seat_fabric1_attribute_value_ids = get_attribute_values(models, seat_main_fabric_id, fabrics_name)
seat_color1_attribute_value_ids = get_attribute_values(models, seat_main_color_id, color_name)
seat_fabric2_attribute_value_ids = get_attribute_values(models, seat_secondary_fabric_id, fabrics_name)
seat_color2_attribute_value_ids = get_attribute_values(models, seat_secondary_color_id, color_name)
seat_fabric1_tcl_attribute_value_ids = get_tcl_attribute_values(models, seat_main_fabric_id)
seat_color1_tcl_attribute_value_ids = get_attribute_values(models, seat_main_color_id, color_name, tcl=True)
seat_fabric2_tcl_attribute_value_ids = get_tcl_attribute_values(models, seat_secondary_fabric_id)
seat_color2_tcl_attribute_value_ids = get_attribute_values(models, seat_secondary_color_id, color_name, tcl=True)
fabric_fabric_attribute_value_ids = get_attribute_values(models, fabric_main_fabric_id, fabrics_name)
fabric_color_attribute_value_ids = get_attribute_values(models, fabric_main_color_id, color_name, tcl=True)


print('Load main and secondary fabrics and colors on product template')
add_pav_on_products(models, product_no_tcl_ids, seat_main_fabric_id, seat_fabric1_attribute_value_ids)
add_pav_on_products(models, product_no_tcl_ids, seat_main_color_id, seat_color1_attribute_value_ids)
add_pav_on_products(models, product_no_tcl_ids, seat_secondary_fabric_id, seat_fabric2_attribute_value_ids)
add_pav_on_products(models, product_no_tcl_ids, seat_secondary_color_id, seat_color2_attribute_value_ids)
add_pav_on_products(models, tcl_ids, seat_main_fabric_id, seat_fabric1_tcl_attribute_value_ids)
add_pav_on_products(models, tcl_ids, seat_main_color_id, seat_color1_tcl_attribute_value_ids)
add_pav_on_products(models, tcl_ids, seat_secondary_fabric_id, seat_fabric2_tcl_attribute_value_ids)
add_pav_on_products(models, tcl_ids, seat_secondary_color_id, seat_color2_tcl_attribute_value_ids)
add_pav_on_products(models, [fabric_product_id], fabric_main_fabric_id, fabric_fabric_attribute_value_ids)
add_pav_on_products(models, [fabric_product_id], fabric_main_color_id, fabric_color_attribute_value_ids)

print('Load exclusion ptav.')
exclusion_to_create = []
i = 0
for color_attribute_value in attribute_combinaison_inverse:
    print("Color %s - %s/%s" % (color_attribute_value, i, len(attribute_combinaison_inverse)))
    exclusion_to_create += configure_ptav_exclusion(color_attribute_value,
                                                    [fabric_main_color_id],
                                                    [fabric_main_fabric_id],
                                                    [fabric_product_id])
    exclusion_to_create += configure_ptav_exclusion(color_attribute_value,
                                                    [seat_main_color_id, seat_secondary_color_id],
                                                    [seat_main_fabric_id, seat_secondary_fabric_id],
                                                    product_no_tcl_ids)
    i += 1
if exclusion_to_create:
    print("Create Missing exclusion ptav")
    exclusion_ids = models.execute_kw(db, uid, password, 'product.template.attribute.exclusion', 'create', [exclusion_to_create])
    print('%s Exclusion PTAV created' % len(exclusion_ids))

print('done')