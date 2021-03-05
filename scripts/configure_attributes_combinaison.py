# -*- coding: utf-8 -*-
"""
Configure attributes combinaison.
"""

import argparse
import logging
from openpyxl import load_workbook


def read_csv(attributes_xlsx):
    attribute_combinaison = {}
    wb = load_workbook(attributes_xlsx)
    sheet = wb.active
    for row in range(2, end_row + 1):
        cell_main_attribute = sheet.cell(row=row, column=1)
        main_attribute_value = cell_main_attribute.value
        attribute_combinaison[main_attribute_value] = []
        column = 2
        cell_secondary_attribute = sheet.cell(row=row, column=column)
        secondary_attribute_value = cell_secondary_attribute.value
        while secondary_attribute_value:
            attribute_combinaison[main_attribute_value].append(str(secondary_attribute_value))
            column += 1
            cell_secondary_attribute = sheet.cell(row=row, column=column)
            secondary_attribute_value = cell_secondary_attribute.value
    return attribute_combinaison


def inverse_attribute_combinaison(attribute_combinaison):
    attribute_combinaison_inverse = {}
    for key, values in attribute_combinaison.items():
        for value in values:
            if value not in attribute_combinaison_inverse:
                attribute_combinaison_inverse[value] = [key]
            else:
                attribute_combinaison_inverse[value].append(key)
    return attribute_combinaison_inverse


def configure_product_attribute_exclusion(exclusions_values):
    k = 0
    logging.info('Start configuring attributes combinaison.')
    for exclusion_id in exclusions_values:
        exclusion = session.env['product.template.attribute.exclusion'].browse(exclusion_id)
        exclusion.write({'value_ids': [(6, 0, exclusions_values[exclusion_id])]})
        k += 1
        if k % 100 == 0:
            logging.info('%s/%s product templates attribute exculsion configured' % (k, len(exclusions_values)))
            session.env.cr.commit()
    session.cr.commit()
    logging.info('%s/%s product templates attribute exculsion configured' % (k, len(exclusions_values)))


parser = argparse.ArgumentParser(description="")
parser.add_argument('database')
parser.add_argument('product_category_id')
parser.add_argument('main_attribute_id')
parser.add_argument('secondary_attribute_id')
parser.add_argument('xlsx_file')
args = parser.parse_args()

end_row = int(input("What is the last line of your xlsx file? "))
session.open(db=args.database)
attribute_combinaison = read_csv(args.xlsx_file)
attribute_combinaison_inverse = inverse_attribute_combinaison(attribute_combinaison)

logging.info('Load product category.')
product_category = session.env['product.category'].search([('id', 'child_of', int(args.product_category_id))])
logging.info('Product category loaded %s.' % product_category.mapped('name'))

logging.info('Load product templates.')
products = session.env['product.template'].search([('categ_id', 'in', product_category.ids)])
logging.info('%s Product templates loaded.' % len(products.ids))

main_attribute_id = int(args.main_attribute_id)
secondary_attribute_id = int(args.secondary_attribute_id)

logging.info('Start configuring attributes combinaison values.')
exclusions_values = {}
i = 0
for product_tmpl in products:
    for secondary_attribute_value in attribute_combinaison_inverse:
        product_template_secondary_attribute = session.env['product.template.attribute.value'].search([('name', '=', secondary_attribute_value), ('product_tmpl_id', '=', product_tmpl.id), ('attribute_id', '=', secondary_attribute_id)])
        allowed_product_template_main_attribute = session.env['product.template.attribute.value'].search([('name', 'in', attribute_combinaison_inverse[secondary_attribute_value]), ('product_tmpl_id', '=', product_tmpl.id), ('attribute_id', '=', main_attribute_id)])
        product_attribute_line = session.env['product.template.attribute.line'].search([('product_tmpl_id', '=', product_tmpl.id), ('attribute_id', '=', main_attribute_id)])
        unallowed_ptav_main_attribute = session.env['product.template.attribute.value'].search([('id', 'not in', allowed_product_template_main_attribute.ids), ('product_tmpl_id', '=', product_tmpl.id), ('attribute_id', '=', main_attribute_id), ('product_attribute_value_id', 'in', product_attribute_line.value_ids.ids)])
        if unallowed_ptav_main_attribute:
            exclusion = product_template_secondary_attribute.exclude_for.filtered(lambda ex: ex.product_tmpl_id.id == product_tmpl.id)
            if not exclusion:
                exclusion = session.env['product.template.attribute.exclusion'].create({'product_template_attribute_value_id': product_template_secondary_attribute.id,'product_tmpl_id': product_tmpl.id})
            # missing_values = unallowed_ptav_main_attribute - exclusion.value_ids
            # if missing_values:
            if exclusions_values.get(exclusion.id, False):
                exclusions_values[exclusion.id].extend(unallowed_ptav_main_attribute.ids)
            else:
                exclusions_values[exclusion.id] = unallowed_ptav_main_attribute.ids
    i += 1
    logging.info('%s product template attribute exclusion values configured' % (i))
    if i % 50 == 0:
        configure_product_attribute_exclusion(exclusions_values)
        exclusions_values = {}
logging.info('%s product template attribute exclusion values configured' % (i))
configure_product_attribute_exclusion(exclusions_values)

session.env.cr.commit()
session.close()
logging.info('Done')

