# -*- coding: utf-8 -*-
"""
Archive unused attribute variants.
"""

import argparse
import logging
from openpyxl import load_workbook


def read_csv(attributes_xlsx):
    attribute_combinaison = {}
    wb = load_workbook(attributes_xlsx)
    sheet = wb.active
    for row in range(2, end_row + 1):
        cell_main_attribute = sheet.cell(row=row, column=1)
        main_attribute = cell_main_attribute.value
        cell_secondary_attribute = sheet.cell(row=row, column=2)
        secondary_attribute = cell_secondary_attribute.value
        attribute_combinaison[main_attribute] = secondary_attribute.split(',')
    return attribute_combinaison


parser = argparse.ArgumentParser(description="")
parser.add_argument('database')
parser.add_argument('product_template_id')
parser.add_argument('main_attribute_id')
parser.add_argument('secondary_attribute_id')
parser.add_argument('xlsx_file')
args = parser.parse_args()

end_row = int(input("What is the last line of your xlsx file? "))
session.open(db=args.database)
attribute_combinaison = read_csv(args.xlsx_file)
product_tmpl = session.env['product.template'].browse(int(args.product_template_id))
main_attribute_id = int(args.main_attribute_id)
secondary_attribute_id = int(args.secondary_attribute_id)

logging.info('Start deleting unnecessary variants.')
product_to_delete = session.env['product.product']
for main_attribute in attribute_combinaison:
    product_template_main_attribute = session.env['product.template.attribute.value'].search([('product_attribute_value_id.name', '=', main_attribute), ('product_tmpl_id', '=', product_tmpl.id), ('attribute_id', '=', main_attribute_id)])
    product_template_secondary_attribute = session.env['product.template.attribute.value'].search([('product_attribute_value_id.name', 'in', attribute_combinaison[main_attribute]), ('product_tmpl_id', '=', product_tmpl.id), ('attribute_id', '=', secondary_attribute_id)])
    for product in session.env['product.product'].search([('product_tmpl_id', '=', product_tmpl.id), ('product_template_attribute_value_ids', 'in', product_template_main_attribute.ids)]):
        owned_pt_secondary_attributes = product.product_template_attribute_value_ids.filtered(lambda ptav: ptav.attribute_id.id == secondary_attribute_id)
        for ptav in owned_pt_secondary_attributes:
            if ptav not in product_template_secondary_attribute:
                product_to_delete |= product
                break
product_to_delete.write({'active': False})
logging.info('Unnecessary product variants deleted')

session.cr.commit()
session.close()
logging.info('Done')

