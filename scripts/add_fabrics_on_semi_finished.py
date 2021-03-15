import xmlrpc.client
import argparse
import logging
from openpyxl import load_workbook


def read_csv(attributes_xlsx):
    attribute_combinaison = {}
    wb = load_workbook(attributes_xlsx)
    sheet = wb.active
    row = 2
    while sheet.cell(row=row, column=1).value:
        cell_fabric_attribute = sheet.cell(row=row, column=1)
        fabric_attribute_value = cell_fabric_attribute.value
        attribute_combinaison[fabric_attribute_value] = []
        column = 2
        cell_color_attribute = sheet.cell(row=row, column=column)
        color_attribute_value = cell_color_attribute.value
        while color_attribute_value:
            attribute_combinaison[fabric_attribute_value].append(str(color_attribute_value))
            column += 1
            cell_color_attribute = sheet.cell(row=row, column=column)
            color_attribute_value = cell_color_attribute.value
        row += 1
    return attribute_combinaison


def inverse_attribute_combinaison(attribute_combinaison):
    attribute_combinaison_inverse = {}
    for key, values in attribute_combinaison.items():
        for value in values:
            if value not in attribute_combinaison_inverse:
                attribute_combinaison_inverse[value] = [key]
            else:
                attribute_combinaison_inverse[value].append(key)
    return attribute_combinaison_inverse


def add_pav_on_products(models, product_ids, attribute_id, attribute_value_ids):
    ptal_ids = models.execute_kw(db, uid, password,
                                 'product.template.attribute.line',
                                 'search',
                                 [[
                                     ('product_tmpl_id', 'in', product_ids),
                                     ('attribute_id', '=', attribute_id),
                                 ]])
    models.execute_kw(db, uid, password, 'product.template.attribute.line', 'write', [ptal_ids, {
        'value_ids': [(6, 0, attribute_value_ids)],
    }])


def get_no_tcl_product_ids(product_ids):
    no_tcl_ids = []
    for product in product_ids:
        if 'c.o.m.' not in product["name"].lower() and 'com' not in product["name"].lower() and 'tcl' not in product["default_code"].lower():
            no_tcl_ids.append(product["id"])
    return no_tcl_ids


def get_attribute_values(models, attribute_id, values_name):
    pav_ids = models.execute_kw(db, uid, password,
                                'product.attribute.value',
                                'search',
                                [[
                                    ('name', 'in', values_name),
                                    ('attribute_id', '=', attribute_id),
                                ]])
    return pav_ids


parser = argparse.ArgumentParser(description="Add Fabrics on semi finished")
parser.add_argument('url')
parser.add_argument('database')
parser.add_argument('user')
parser.add_argument('password')
parser.add_argument('product_category_id')
parser.add_argument('main_fabric_attribute_id')
parser.add_argument('main_color_attribute_id')
parser.add_argument('secondary_fabric_attribute_id')
parser.add_argument('secondary_color_attribute_id')
parser.add_argument('xlsx_file')
args = parser.parse_args()

db =args.database
password = args.password
main_fabric_id = int(args.main_fabric_attribute_id)
main_color_id = int(args.main_color_attribute_id)
secondary_fabric_id = int(args.secondary_fabric_attribute_id)
secondary_color_id = int(args.secondary_color_attribute_id)

common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(args.url))
uid = common.authenticate(db, args.user, password, {})
models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(args.url))

print('Load product category.')
product_category_ids = models.execute_kw(db, uid, password, 'product.category', 'search', [[('id', 'child_of', int(args.product_category_id))]])
print('%s Product category loaded.' % len(product_category_ids))

print('Load product templates.')
products_info = models.execute_kw(db, uid, password, 'product.template', 'search_read', [[('categ_id', 'in', product_category_ids)]], {'fields': ['name', 'default_code']})
print('%s Product templates loaded.' % len(products_info))

print('Read xlsx.')
attribute_combinaison = read_csv(args.xlsx_file)
attribute_combinaison_inverse = inverse_attribute_combinaison(attribute_combinaison)

print('Set Data1.')
fabrics_name = list(attribute_combinaison.keys())
color_name = list(attribute_combinaison_inverse.keys())
product_no_tcl_ids = get_no_tcl_product_ids(products_info)
fabric1_attribute_value_ids = get_attribute_values(models, main_fabric_id, fabrics_name)
color1_attribute_value_ids = get_attribute_values(models, main_color_id, color_name)
fabric2_attribute_value_ids = get_attribute_values(models, secondary_fabric_id, fabrics_name)
color2_attribute_value_ids = get_attribute_values(models, secondary_color_id, color_name)

print('Load main and secondary fabrics and colors on product template')
add_pav_on_products(models, product_no_tcl_ids, main_fabric_id, fabric1_attribute_value_ids)
add_pav_on_products(models, product_no_tcl_ids, main_color_id, color1_attribute_value_ids)
add_pav_on_products(models, product_no_tcl_ids, secondary_fabric_id, fabric2_attribute_value_ids)
add_pav_on_products(models, product_no_tcl_ids, secondary_color_id, color2_attribute_value_ids)

print('Load exclusion ptav.')
exclusion_to_create = []
for product_id in product_no_tcl_ids:
    for secondary_attribute_value in attribute_combinaison_inverse:
        for attribute_color_id, attribute_fabric_id in zip([main_color_id, secondary_color_id], [main_fabric_id, secondary_fabric_id]):
            ptav_color_id = models.execute_kw(db, uid, password,
                                                   'product.template.attribute.value',
                                                   'search',
                                                   [[
                                                       ('name', '=', secondary_attribute_value),
                                                       ('attribute_id', '=', attribute_color_id),
                                                       ('product_tmpl_id', '=', product_id)
                                                   ]]
                                                   )
            unallowed_ptav_fabric_ids = models.execute_kw(db, uid, password,
                                                               'product.template.attribute.value',
                                                               'search',
                                                               [[
                                                                   ('name', 'not in', attribute_combinaison_inverse[secondary_attribute_value]),
                                                                   ('product_tmpl_id', '=', product_id),
                                                                   ('attribute_id', '=', attribute_fabric_id)
                                                               ]]
                                                               )
            if ptav_color_id and unallowed_ptav_fabric_ids:
                ptae_id = models.execute_kw(db, uid, password,
                                            'product.template.attribute.exclusion',
                                            'search',
                                            [[
                                                ('product_template_attribute_value_id', '=', ptav_color_id),
                                                ('product_tmpl_id', '=', product_id),
                                            ]]
                                            )
                if ptae_id:
                    models.execute_kw(db, uid, password, 'product.template.attribute.exclusion', 'write', [ptae_id, {
                        'value_ids': [(6, 0, unallowed_ptav_fabric_ids)],
                    }])
                else:
                    exclusion_to_create.append({'product_template_attribute_value_id': ptav_color_id,
                                                'product_tmpl_id': product_id,
                                                'value_ids': [(6, 0, unallowed_ptav_fabric_ids)]})
if exclusion_to_create:
    exclusion_ids = models.execute_kw(db, uid, password, 'product.template.attribute.exclusion', 'create', exclusion_to_create)
    print('%s Exclusion PTAV created' % len(exclusion_ids))

print('done')